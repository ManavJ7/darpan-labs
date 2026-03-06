import csv
import os
import re
from fpdf import FPDF
import openpyxl

TRAINING_DIR = "training"
OUTPUT_FILE = "Primary Survey Responses.pdf"


def extract_participant_number(filename):
    match = re.search(r"interview_transcript_(\d+)", filename)
    return int(match.group(1)) if match else 0


def read_transcript_csv(filepath):
    qa_pairs = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            question = row.get("question_text", "").strip()
            answer = row.get("answer_text", "").strip()
            if question:
                qa_pairs.append((question, answer))
    return qa_pairs


def read_transcript_xlsx(filepath):
    qa_pairs = []
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return qa_pairs
    # Skip header row
    for row in rows[1:]:
        question = str(row[1] or "").strip()
        answer = str(row[2] or "").strip()
        if question:
            qa_pairs.append((question, answer))
    return qa_pairs


def sanitize(text):
    """Replace characters that latin-1 can't encode."""
    replacements = {
        "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"',
        "\u2013": "-", "\u2014": "-", "\u2026": "...", "\u00a0": " ",
        "\u2192": "->", "\u2190": "<-", "\ufeff": "",
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text.encode("latin-1", "replace").decode("latin-1")


def build_pdf():
    # Collect all participants: xlsx files (1, 2) + csv files (3-18)
    xlsx_mapping = {
        "training_data_1.xlsx": 1,
        "training_data_2.xlsx": 2,
    }
    participants = []  # list of (participant_num, filepath)

    for fname, pnum in sorted(xlsx_mapping.items(), key=lambda x: x[1]):
        fpath = os.path.join(TRAINING_DIR, fname)
        if os.path.exists(fpath):
            participants.append((pnum, fpath, "xlsx"))

    csv_files = sorted(
        [f for f in os.listdir(TRAINING_DIR) if f.endswith(".csv")],
        key=extract_participant_number,
    )
    for fname in csv_files:
        pnum = extract_participant_number(fname)
        participants.append((pnum, os.path.join(TRAINING_DIR, fname), "csv"))

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # Title page
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 28)
    pdf.ln(80)
    pdf.cell(0, 15, "Primary Survey Responses", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 14)
    pdf.ln(10)
    pdf.cell(0, 10, f"{len(participants)} Participants", align="C", new_x="LMARGIN", new_y="NEXT")

    for participant_num, filepath, filetype in participants:
        if filetype == "xlsx":
            qa_pairs = read_transcript_xlsx(filepath)
        else:
            qa_pairs = read_transcript_csv(filepath)

        if not qa_pairs:
            continue

        # New page for each participant
        pdf.add_page()

        # Participant header
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_fill_color(41, 65, 122)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 12, f"  Participant {participant_num}", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(6)

        for q_idx, (question, answer) in enumerate(qa_pairs, 1):
            question = sanitize(question)
            answer = sanitize(answer)

            # Check if we need a new page (rough estimate)
            if pdf.get_y() > 250:
                pdf.add_page()

            # Question
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 6, f"Q{q_idx}. {question}", new_x="LMARGIN", new_y="NEXT")

            # Answer
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(60, 60, 60)
            pdf.multi_cell(0, 6, f"A: {answer if answer else '[No response]'}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(4)

    pdf.output(OUTPUT_FILE)
    print(f"Generated '{OUTPUT_FILE}' with {len(participants)} participant transcripts.")


if __name__ == "__main__":
    build_pdf()
