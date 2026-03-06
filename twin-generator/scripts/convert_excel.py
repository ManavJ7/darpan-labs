"""
Convert training files (xlsx + csv) to the JSON format expected by the pipeline.

Scans training/ folder for:
  - training_data_*.xlsx  (P01, P02, ...)
  - interview_transcript_*.csv  (P03, P04, ...)

Each file = 1 participant with columns: module_id, question_text, answer_text.

Outputs:
  - data/input/real_qa_pairs.json  (combined, all participants)
  - data/output/P{NN}/real_qa_pairs.json  (per-participant copy)

Usage:
    python scripts/convert_excel.py
"""
import csv
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import TRAINING_DIR, INPUT_DIR, participant_output_dir

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def find_training_files() -> list[tuple[Path, str]]:
    """Find all training files and map each to a participant ID.

    Returns sorted list of (path, participant_id) tuples.
    """
    if not TRAINING_DIR.exists():
        raise FileNotFoundError(f"Training directory not found: {TRAINING_DIR}")

    results: list[tuple[Path, str, int]] = []

    # xlsx files: training_data_1.xlsx -> P01
    for p in TRAINING_DIR.glob("training_data_*.xlsx"):
        m = re.search(r"(\d+)", p.stem.replace("training_data_", ""))
        if m:
            num = int(m.group(1))
            results.append((p, f"P{num:02d}", num))

    # csv files: interview_transcript_3.csv -> P03
    for p in TRAINING_DIR.glob("interview_transcript_*.csv"):
        m = re.search(r"(\d+)", p.stem.replace("interview_transcript_", ""))
        if m:
            num = int(m.group(1))
            results.append((p, f"P{num:02d}", num))

    if not results:
        raise FileNotFoundError(f"No training files found in {TRAINING_DIR}")

    results.sort(key=lambda x: x[2])
    return [(path, pid) for path, pid, _ in results]


def _extract_col_map(headers: list[str]) -> dict[str, int]:
    """Map column names to indices from a list of lowercase header strings."""
    col_map = {}
    for i, h in enumerate(headers):
        if "module" in h:
            col_map["module_id"] = i
        elif "question" in h:
            col_map["question_text"] = i
        elif "answer" in h:
            col_map["answer_text"] = i
    if len(col_map) < 3:
        raise ValueError(f"Expected columns module_id, question_text, answer_text. Found headers: {headers}")
    return col_map


def read_excel(path: Path) -> list[dict]:
    """Read a training Excel file and return list of QA dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = _extract_col_map(headers)

    qa_pairs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        module_id = str(row_list[col_map["module_id"]] or "").strip()
        question_text = str(row_list[col_map["question_text"]] or "").strip()
        answer_text = str(row_list[col_map["answer_text"]] or "").strip()

        if question_text and answer_text:
            qa_pairs.append({
                "module_id": module_id,
                "question_text": question_text,
                "answer_text": answer_text,
            })

    wb.close()
    return qa_pairs


def read_csv(path: Path) -> list[dict]:
    """Read a training CSV file and return list of QA dicts."""
    rows = None
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with open(path, encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                fieldnames = [n.strip().lower() for n in (reader.fieldnames or [])]
                col_map = _extract_col_map(fieldnames)
                # Map original fieldnames to our keys
                orig_fields = reader.fieldnames or []
                field_for = {
                    "module_id": orig_fields[col_map["module_id"]],
                    "question_text": orig_fields[col_map["question_text"]],
                    "answer_text": orig_fields[col_map["answer_text"]],
                }
                rows = []
                for row in reader:
                    module_id = (row.get(field_for["module_id"]) or "").strip()
                    question_text = (row.get(field_for["question_text"]) or "").strip()
                    answer_text = (row.get(field_for["answer_text"]) or "").strip()
                    if question_text and answer_text:
                        rows.append({
                            "module_id": module_id,
                            "question_text": question_text,
                            "answer_text": answer_text,
                        })
            break
        except UnicodeDecodeError:
            continue

    if rows is None:
        raise ValueError(f"Could not read {path} with any supported encoding")
    return rows


def main():
    files = find_training_files()
    print(f"Found {len(files)} training files in {TRAINING_DIR}")

    all_participants = []

    for path, pid in files:
        if path.suffix == ".csv":
            qa_pairs = read_csv(path)
        else:
            qa_pairs = read_excel(path)
        print(f"  {pid}: {path.name} -> {len(qa_pairs)} Q&A pairs")

        participant = {
            "participant_id": pid,
            "source_file": path.name,
            "qa_pairs": qa_pairs,
        }
        all_participants.append(participant)

        # Per-participant copy
        out_dir = participant_output_dir(pid)
        with open(out_dir / "real_qa_pairs.json", "w") as f:
            json.dump([participant], f, indent=2)

    # Combined output
    combined_path = INPUT_DIR / "real_qa_pairs.json"
    with open(combined_path, "w") as f:
        json.dump(all_participants, f, indent=2)
    print(f"\nCombined output: {combined_path} ({len(all_participants)} participants)")


if __name__ == "__main__":
    main()
