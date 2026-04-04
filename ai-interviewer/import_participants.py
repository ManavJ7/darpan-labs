"""
Import 17 participants' interview data from Excel into the AI Interviewer database.

- Creates user records for each participant (P01-P17)
- Creates interview sessions with all 8 modules marked complete
- Creates interview turns for every Q&A pair
- For MCQ/select questions: maps answer text to option number + value
- Stores both the answer text AND option_number in answer_structured JSONB
"""

import json
import os
import sys
import uuid
from datetime import datetime, timezone, timedelta

import openpyxl
import psycopg2
from psycopg2.extras import Json

# ─── Config ─────────────────────────────────────────────────────────────
EXCEL_PATH = "/Users/manavrsjain/Documents/darpan-labs-V2/twin-generator/data/Interview_Training_Data_18_Participants.xlsx"
SEED_DIR = "/Users/manavrsjain/Desktop/darpan-labs-V2-reorganized/ai-interviewer/backend/seed_data/question_banks"
DB_URL = "postgresql://manavrsjain@localhost:5432/darpan"


# ─── 1. Load question banks ────────────────────────────────────────────
def load_question_banks():
    """Load all question banks and build lookup structures."""
    all_questions = {}  # bank_qid -> full question object
    module_questions = {}  # module_id -> [questions in order]

    for fname in sorted(os.listdir(SEED_DIR)):
        if not fname.endswith('.json'):
            continue
        with open(os.path.join(SEED_DIR, fname)) as f:
            bank = json.load(f)

        module_id = bank['module_id']
        module_questions[module_id] = []

        for q in bank['questions']:
            q['_module'] = module_id
            q['_options_lookup'] = {}
            # Build options lookup: lowercase label -> (option_number, value)
            for idx, opt in enumerate(q.get('options', []), 1):
                label = opt['label'].lower().strip()
                value = opt['value']
                q['_options_lookup'][label] = (idx, value)
            all_questions[q['question_id']] = q
            module_questions[module_id].append(q)

    return all_questions, module_questions


# ─── 2. Parse Excel data ───────────────────────────────────────────────
def parse_excel():
    """Parse the pivot sheet to get all Q&A data per participant."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # --- Parse pivot sheet (canonical questions) ---
    ws = wb['Responses (Pivot)']
    header_row = 4

    # Get participant columns
    participants = []
    participant_cols = {}  # participant_id -> column_number
    for c in range(5, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            pid = str(val).strip()
            participants.append(pid)
            participant_cols[pid] = c

    # Parse questions and answers
    questions_data = []  # [{qid, module, question_text, answers: {P01: "...", P02: "..."}}]
    current_module = None
    for r in range(5, ws.max_row + 1):
        row_label = str(ws.cell(row=r, column=1).value or '').strip()
        qid = str(ws.cell(row=r, column=2).value or '').strip()
        module = str(ws.cell(row=r, column=3).value or '').strip()
        question_text = str(ws.cell(row=r, column=4).value or '').strip()

        if module:
            current_module = module

        if not qid or not current_module:
            continue

        answers = {}
        for pid in participants:
            col = participant_cols[pid]
            ans = ws.cell(row=r, column=col).value
            if ans is not None:
                answers[pid] = str(ans).strip()

        questions_data.append({
            'excel_qid': qid,
            'module': current_module,
            'question_text': question_text,
            'answers': answers,
        })

    # --- Parse long format for follow-up questions ---
    ws_long = wb['All Responses (Long)']
    followups = {}  # participant -> [(module, qid, question, answer)]
    current_participant = None
    for r in range(5, ws_long.max_row + 1):
        participant = ws_long.cell(row=r, column=2).value
        module = ws_long.cell(row=r, column=3).value
        qid = ws_long.cell(row=r, column=4).value
        question = ws_long.cell(row=r, column=5).value
        answer = ws_long.cell(row=r, column=6).value
        qtype = ws_long.cell(row=r, column=7).value

        if participant:
            current_participant = str(participant).strip()

        if qtype and str(qtype).strip() == 'Follow-up' and current_participant and answer:
            if current_participant not in followups:
                followups[current_participant] = []
            followups[current_participant].append({
                'module': str(module).strip() if module else '',
                'qid': str(qid).strip() if qid else 'followup',
                'question': str(question).strip() if question else '',
                'answer': str(answer).strip(),
            })

    # --- Parse participant profiles ---
    ws_prof = wb['Participant Profiles']
    profiles = {}  # pid -> {attribute: value}
    # Row 4 is header: '', 'Attribute', P01, P02, ...
    prof_cols = {}
    for c in range(3, ws_prof.max_column + 1):
        val = ws_prof.cell(row=4, column=c).value
        if val:
            prof_cols[str(val).strip()] = c

    for r in range(5, ws_prof.max_row + 1):
        attr = ws_prof.cell(row=r, column=2).value
        if not attr:
            continue
        attr = str(attr).strip()
        for pid, col in prof_cols.items():
            val = ws_prof.cell(row=r, column=col).value
            if val:
                if pid not in profiles:
                    profiles[pid] = {}
                profiles[pid][attr] = str(val).strip()

    return participants, questions_data, followups, profiles


# ─── 3. Map Excel Q IDs to question bank IDs ───────────────────────────
def build_question_mapping(questions_data, all_questions, module_questions):
    """Map Excel Q IDs to question bank IDs using text matching."""
    mapping = {}  # excel_qid -> bank_qid

    for qd in questions_data:
        excel_qid = qd['excel_qid']
        module = qd['module']
        q_text = qd['question_text'].lower()

        # M8 questions already have the right format
        if excel_qid.startswith('M8_'):
            if excel_qid in all_questions:
                mapping[excel_qid] = excel_qid
            continue

        # For M1-M7, match by question text prefix
        best_match = None
        best_score = 0
        for bank_q in module_questions.get(module, []):
            bank_text = bank_q['question_text'].lower()
            # Compare first 30 chars
            common = 0
            for a, b in zip(q_text, bank_text):
                if a == b:
                    common += 1
                else:
                    break
            if common > best_score:
                best_score = common
                best_match = bank_q['question_id']

        if best_match and best_score >= 15:
            mapping[excel_qid] = best_match
        else:
            # Fallback: try containing match
            for bank_q in module_questions.get(module, []):
                bank_text = bank_q['question_text'].lower()
                if q_text[:25] in bank_text or bank_text[:25] in q_text:
                    mapping[excel_qid] = bank_q['question_id']
                    break

    return mapping


# ─── 4. Match MCQ answers to option numbers ────────────────────────────
def match_answer_to_option(answer_text, question):
    """
    For MCQ questions, match the answer text to the closest option.
    Returns: {selected_options: [{option_number, label, value}], raw_answer}
    """
    options = question.get('options', [])
    q_type = question.get('question_type', '')

    if not options or q_type in ('open_text', 'numeric'):
        return None

    answer_lower = answer_text.lower().strip()
    matched = []

    if q_type in ('single_select', 'scale', 'scale_open'):
        # Find best matching single option
        best_idx = None
        best_score = 0
        for idx, opt in enumerate(options, 1):
            label = opt['label'].lower()
            value = opt['value'].lower()

            # Exact match
            if answer_lower == label or answer_lower == value:
                best_idx = idx
                best_score = 1000
                break

            # Check if answer contains the label or vice versa
            score = 0
            if label in answer_lower:
                score = len(label)
            elif answer_lower in label:
                score = len(answer_lower)
            # Also check value
            if value.replace('_', ' ') in answer_lower:
                score = max(score, len(value))

            # Special: check key words
            label_words = set(label.split())
            answer_words = set(answer_lower.split())
            overlap = label_words & answer_words
            if len(overlap) >= 2:
                score = max(score, len(overlap) * 10)

            if score > best_score:
                best_score = score
                best_idx = idx

        if best_idx and best_score > 0:
            opt = options[best_idx - 1]
            matched.append({
                'option_number': best_idx,
                'label': opt['label'],
                'value': opt['value'],
            })

    elif q_type in ('multi_select',):
        # Match multiple options
        for idx, opt in enumerate(options, 1):
            label = opt['label'].lower()
            value = opt['value'].lower()
            # Check if this option appears in the answer
            if (label in answer_lower or
                value.replace('_', ' ') in answer_lower or
                any(w in answer_lower for w in label.split() if len(w) > 4)):
                matched.append({
                    'option_number': idx,
                    'label': opt['label'],
                    'value': opt['value'],
                })

    elif q_type in ('rank_order',):
        # For rank order, try to parse the ranking from the answer
        # Answers often look like "Price > Brand > Ingredients > ..."
        # or "1. Price, 2. Brand, 3. Ingredients"
        parts = []
        # Try splitting by >, then by numbered list, then by commas
        if '>' in answer_text:
            parts = [p.strip() for p in answer_text.split('>')]
        elif any(f'{i}.' in answer_text for i in range(1, 10)):
            import re
            parts = re.split(r'\d+[\.\)]\s*', answer_text)
            parts = [p.strip().rstrip(',').strip() for p in parts if p.strip()]
        else:
            parts = [p.strip() for p in answer_text.split(',')]

        for rank, part in enumerate(parts, 1):
            part_lower = part.lower().strip()
            best_opt = None
            best_score = 0
            for idx, opt in enumerate(options, 1):
                label = opt['label'].lower()
                value = opt['value'].lower()
                score = 0
                if part_lower == label or part_lower == value:
                    score = 1000
                elif label in part_lower or part_lower in label:
                    score = len(min(label, part_lower, key=len))
                elif value.replace('_', ' ') in part_lower:
                    score = len(value)
                # Word overlap
                label_words = set(label.split())
                part_words = set(part_lower.split())
                overlap = label_words & part_words
                if len(overlap) >= 1 and any(len(w) > 3 for w in overlap):
                    score = max(score, len(overlap) * 10)
                if score > best_score:
                    best_score = score
                    best_opt = (idx, opt)

            if best_opt and best_score > 0:
                matched.append({
                    'option_number': best_opt[0],
                    'label': best_opt[1]['label'],
                    'value': best_opt[1]['value'],
                    'rank': rank,
                })

    elif q_type in ('matrix_scale', 'matrix_premium'):
        # Matrix questions - answer is usually a complex string
        # Store as-is with the question context
        return {
            'question_type': q_type,
            'raw_answer': answer_text,
            'note': 'Matrix response - parsed as text',
        }

    if matched:
        result = {
            'question_type': q_type,
            'selected_options': matched,
            'raw_answer': answer_text,
        }
        return result

    # No match found - return raw
    return {
        'question_type': q_type,
        'raw_answer': answer_text,
        'match_failed': True,
        'note': 'Could not match answer to any option',
    }


# ─── 5. Insert into database ───────────────────────────────────────────
def import_data():
    """Main import function."""
    print("Loading question banks...")
    all_questions, module_questions = load_question_banks()
    print(f"  Loaded {len(all_questions)} questions across {len(module_questions)} modules")

    print("Parsing Excel data...")
    participants, questions_data, followups, profiles = parse_excel()
    print(f"  Found {len(participants)} participants, {len(questions_data)} canonical questions")
    print(f"  Follow-ups: {sum(len(v) for v in followups.values())} total across {len(followups)} participants")

    print("Building question ID mapping...")
    mapping = build_question_mapping(questions_data, all_questions, module_questions)
    print(f"  Mapped {len(mapping)}/{len(questions_data)} questions")

    # Show unmapped
    unmapped = [qd['excel_qid'] for qd in questions_data if qd['excel_qid'] not in mapping]
    if unmapped:
        print(f"  WARNING: Unmapped questions: {unmapped}")

    print("\nConnecting to database...")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    try:
        # Check existing data
        cur.execute("SELECT COUNT(*) FROM users WHERE email LIKE '%@darpan-participant.local'")
        existing = cur.fetchone()[0]
        if existing > 0:
            print(f"  Found {existing} existing participant users. Cleaning up...")
            cur.execute("""
                DELETE FROM interview_turns WHERE session_id IN (
                    SELECT id FROM interview_sessions WHERE user_id IN (
                        SELECT id FROM users WHERE email LIKE '%@darpan-participant.local'
                    )
                )
            """)
            cur.execute("""
                DELETE FROM interview_modules WHERE session_id IN (
                    SELECT id FROM interview_sessions WHERE user_id IN (
                        SELECT id FROM users WHERE email LIKE '%@darpan-participant.local'
                    )
                )
            """)
            cur.execute("""
                DELETE FROM interview_sessions WHERE user_id IN (
                    SELECT id FROM users WHERE email LIKE '%@darpan-participant.local'
                )
            """)
            cur.execute("DELETE FROM users WHERE email LIKE '%@darpan-participant.local'")
            conn.commit()
            print("  Cleaned up existing data.")

        modules_list = ['M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8']

        total_turns = 0
        total_matched_options = 0
        total_mcq = 0

        for pid in participants:
            print(f"\n  Importing {pid}...")

            # Get profile info
            prof = profiles.get(pid, {})
            display_name = f"Participant {pid}"
            location = prof.get('Location', '')
            occupation = prof.get('Occupation', '')

            # Determine sex/age from profile if available
            sex = prof.get('Sex', None)
            age_str = prof.get('Age', None)
            age = None
            if age_str:
                try:
                    age = int(age_str)
                except (ValueError, TypeError):
                    pass

            # 5a. Create user
            user_id = uuid.uuid4()
            email = f"{pid.lower()}@darpan-participant.local"
            cur.execute("""
                INSERT INTO users (id, email, display_name, auth_provider_id, sex, age, profile_completed, is_admin, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, true, false, NOW(), NOW())
            """, (str(user_id), email, display_name, f"imported_{pid}", sex, age))

            # 5b. Create interview session
            session_id = uuid.uuid4()
            base_time = datetime(2026, 3, 15, 10, 0, 0, tzinfo=timezone.utc)
            cur.execute("""
                INSERT INTO interview_sessions (id, user_id, status, input_mode, language_preference, started_at, ended_at, total_duration_sec, settings)
                VALUES (%s, %s, 'completed', 'text', 'en', %s, %s, %s, %s)
            """, (
                str(session_id), str(user_id),
                base_time,
                base_time + timedelta(hours=2),
                7200,
                Json({'source': 'excel_import', 'participant_id': pid}),
            ))

            # 5c. Create interview modules (all completed)
            module_turn_counts = {m: 0 for m in modules_list}
            module_ids = {}
            for mod in modules_list:
                mod_id = uuid.uuid4()
                module_ids[mod] = mod_id
                cur.execute("""
                    INSERT INTO interview_modules (id, session_id, module_id, status, started_at, ended_at, question_count, coverage_score, confidence_score, signals_captured, completion_eval)
                    VALUES (%s, %s, %s, 'completed', %s, %s, 0, 1.0, 1.0, %s, %s)
                """, (
                    str(mod_id), str(session_id), mod,
                    base_time, base_time + timedelta(minutes=15),
                    Json([]), Json({'imported': True, 'source': 'excel'}),
                ))

            # 5d. Create interview turns for canonical questions
            turn_index = 0
            for qd in questions_data:
                excel_qid = qd['excel_qid']
                module = qd['module']
                question_text = qd['question_text']
                answer = qd['answers'].get(pid)

                if not answer:
                    continue

                bank_qid = mapping.get(excel_qid)
                bank_question = all_questions.get(bank_qid) if bank_qid else None

                # Build question meta
                question_meta = {
                    'excel_qid': excel_qid,
                    'bank_qid': bank_qid,
                    'question_type': bank_question['question_type'] if bank_question else 'open_text',
                }

                # Build answer_structured with option matching
                answer_structured = None
                if bank_question and bank_question.get('options'):
                    total_mcq += 1
                    matched = match_answer_to_option(answer, bank_question)
                    if matched:
                        answer_structured = matched
                        if matched.get('selected_options'):
                            total_matched_options += 1

                turn_id = uuid.uuid4()
                cur.execute("""
                    INSERT INTO interview_turns (id, session_id, module_id, turn_index, role, input_mode, question_text, question_meta, answer_text, answer_language, answer_structured, created_at)
                    VALUES (%s, %s, %s, %s, 'user', 'text', %s, %s, %s, 'en', %s, %s)
                """, (
                    str(turn_id), str(session_id), module,
                    turn_index, question_text,
                    Json(question_meta),
                    answer,
                    Json(answer_structured) if answer_structured else None,
                    base_time + timedelta(seconds=turn_index * 30),
                ))
                turn_index += 1
                module_turn_counts[module] = module_turn_counts.get(module, 0) + 1
                total_turns += 1

            # 5e. Create interview turns for follow-up questions
            pid_followups = followups.get(pid, [])
            for fu in pid_followups:
                turn_id = uuid.uuid4()
                question_meta = {
                    'excel_qid': fu['qid'],
                    'bank_qid': None,
                    'question_type': 'follow_up',
                    'is_followup': True,
                }
                cur.execute("""
                    INSERT INTO interview_turns (id, session_id, module_id, turn_index, role, input_mode, question_text, question_meta, answer_text, answer_language, answer_structured, created_at)
                    VALUES (%s, %s, %s, %s, 'user', 'text', %s, %s, %s, 'en', %s, %s)
                """, (
                    str(turn_id), str(session_id), fu['module'],
                    turn_index, fu['question'],
                    Json(question_meta),
                    fu['answer'],
                    Json({'question_type': 'follow_up', 'raw_answer': fu['answer']}),
                    base_time + timedelta(seconds=turn_index * 30),
                ))
                turn_index += 1
                module_turn_counts[fu['module']] = module_turn_counts.get(fu['module'], 0) + 1
                total_turns += 1

            # 5f. Update module question counts
            for mod, count in module_turn_counts.items():
                if mod in module_ids:
                    cur.execute("""
                        UPDATE interview_modules SET question_count = %s WHERE id = %s
                    """, (count, str(module_ids[mod])))

            print(f"    User: {email}")
            print(f"    Turns: {turn_index} ({sum(1 for qd in questions_data if pid in qd['answers'])} canonical + {len(pid_followups)} follow-ups)")

        conn.commit()
        print(f"\n{'='*60}")
        print(f"IMPORT COMPLETE")
        print(f"  Participants: {len(participants)}")
        print(f"  Total turns: {total_turns}")
        print(f"  MCQ questions: {total_mcq} total, {total_matched_options} matched to options")
        print(f"  Match rate: {total_matched_options/total_mcq*100:.1f}%" if total_mcq > 0 else "  No MCQ")
        print(f"{'='*60}")

        # Verification
        cur.execute("SELECT COUNT(*) FROM users WHERE email LIKE '%@darpan-participant.local'")
        print(f"\nVerification:")
        print(f"  Users in DB: {cur.fetchone()[0]}")
        cur.execute("SELECT COUNT(*) FROM interview_sessions WHERE settings->>'source' = 'excel_import'")
        print(f"  Sessions in DB: {cur.fetchone()[0]}")
        cur.execute("""
            SELECT COUNT(*) FROM interview_turns WHERE session_id IN (
                SELECT id FROM interview_sessions WHERE settings->>'source' = 'excel_import'
            )
        """)
        print(f"  Turns in DB: {cur.fetchone()[0]}")

        # Show a sample MCQ match
        print("\nSample MCQ matches (P01):")
        cur.execute("""
            SELECT question_text, answer_text, answer_structured
            FROM interview_turns
            WHERE session_id IN (
                SELECT id FROM interview_sessions WHERE user_id IN (
                    SELECT id FROM users WHERE email = 'p01@darpan-participant.local'
                )
            )
            AND answer_structured IS NOT NULL
            AND answer_structured->>'selected_options' IS NOT NULL
            LIMIT 5
        """)
        for row in cur.fetchall():
            q = row[0][:50]
            a = row[1][:40]
            s = row[2]
            if isinstance(s, str):
                s = json.loads(s)
            opts = s.get('selected_options', [])
            if opts:
                opt_str = ', '.join(f"Opt#{o['option_number']}: {o['label'][:25]}" for o in opts)
                print(f"  Q: {q}...")
                print(f"  A: {a}")
                print(f"  -> {opt_str}")
                print()

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    import_data()
